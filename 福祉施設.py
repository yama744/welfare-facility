#!/usr/bin/env python
# coding: utf-8

# In[1]:


from bs4 import BeautifulSoup
import urllib.request as req
from selenium import webdriver
import time
import pandas as pd
#import chromedriver_binary
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import TimeoutException
import math
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pandas as pd
#import chromedriver_binary
import openpyxl
import glob
import xlrd
import pprint
from collections import OrderedDict
import csv
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import chromedriver_binary
from pyvirtualdisplay import Display
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.keys import Keys


# In[2]:


option = Options()# オプションを用意
#option.add_argument('--user-agent=hogehoge')
option.add_argument('--headless')           # ヘッドレスモードの設定を付与
option.add_argument('--disable-extensions')       # すべての拡張機能を無効にする。ユーザースクリプトも無効にする
option.add_argument('--proxy-server="direct://"') # Proxy経由ではなく直接接続する
option.add_argument('--proxy-bypass-list=*')      # すべてのホスト名
#option.add_argument('--start-maximized')          # 起動時にウィンドウを最大化する
#driver = webdriver.Chrome(options=option)   # Chromeを準備(optionでヘッドレスモードにしている）


# In[16]:


#driver = webdriver.Chrome(options=option) 
driver = webdriver.Chrome() 
url_login=('https://www.wam.go.jp/sfkohyoout/COP000100E0000.do')
time.sleep(5)
#url_login=('https://www.amazon.co.jp/s?bbn=2189188051&rh=n%3A2189188051%2Cp_n_date%3A5871621051&dc&qid=1613727694&rnid=82820051&ref=lp_2189188051_nr_p_n_date_7')
driver.get(url_login)


# In[17]:


#検索
element=driver.find_element_by_xpath('/html/body/div/form/div/table/tbody/tr/td[1]/input[1]')
time.sleep(1)
element.send_keys('大阪府')
driver.find_element_by_xpath('/html/body/div/form/div/table/tbody/tr/td[1]/input[2]').click()
time.sleep(5)
""""#大分類
driver.find_element_by_xpath('/html/body/div[4]/div/form/div[2]/table/tbody/tr[1]/td/div/div/select/option[9]').click()
time.sleep(1)
#小分類
driver.find_element_by_xpath('/html/body/div[4]/div/form/div[2]/table/tbody/tr[2]/td/div/div/select/option[9]').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/div[4]/div/form/div[3]/button[2]').click()
time.sleep(1)
driver.find_element_by_xpath('/html/body/div/div/form/table[2]/tbody/tr[9]/td/div/input[2]').click()
time.sleep(1)"""


# In[18]:


#メインプログラム
#driver.find_element_by_xpath('/html/body/div/div/form/div[6]/div[2]/div[2]/div/select/option[3]').click()

syameilists=[]
yubinlists=[]
jusyolists=[]
tellists=[]
faxlists=[]
urllists=[]
syatyolists=[]
tantolists=[]
#途中から開始
""""for x in range(0,32):
    page=driver.find_element_by_class_name('m05').text
    print(page)
    elements = driver.find_element_by_css_selector('#ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]')
        
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(9) > input[type=submit]
        #driver.scrollTo(elements.location_once_scrolled_into_view['y'])
    loc = elements.location
    x, y = loc['x'], loc['y']
    actions = ActionChains(driver)
        #actions.move_by_offset(1, 1)
    webdriver.ActionChains(driver).move_to_element(elements).perform()
        #actions.click()
        #actions.perform()
    time.sleep(1)
    driver.find_element_by_css_selector('#ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]').click()
        #driver.find_element_by_name("fwListNaviBtnNext").click()
        #driver.find_element_by_xpath('/html/body/div/div/form/div[6]/div[3]/ul/li[8]/input').click()"""
    #time.sleep(1.5)
i=0
while i!=100:
    try:
        #WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'ID_dispDetailBtn')))
        browser_from=driver.find_elements_by_class_name('detail-button')
    except TimeoutException:
        time.sleep(10)
    # ウィンドウハンドルを取得する
        handle_array = driver.window_handles
    # seleniumで操作可能なdriverを切り替える
        driver.switch_to.window(handle_array[0])
        time.sleep(3)
        browser_from=driver.find_elements_by_id('ID_dispDetailBtn')
    try: 
        browser_from[i].click()
    except IndexError:
        page=driver.find_element_by_class_name('m05').text
        print(page)
        elements = driver.find_element_by_css_selector('#ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(7) > input[type=submit]')
        
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(8) > input[type=submit]
        #ID_form_1 > div:nth-child(12) > div:nth-child(3) > ul > li:nth-child(9) > input[type=submit]
        #driver.scrollTo(elements.location_once_scrolled_into_view['y'])
        loc = elements.location
        x, y = loc['x'], loc['y']
        actions = ActionChains(driver)
        #actions.move_by_offset(1, 1)
        webdriver.ActionChains(driver).move_to_element(elements).perform()
        #actions.click()
        #actions.perform()
        time.sleep(1)
        driver.find_element_by_name("fwListNaviBtnNext").click()
        #driver.find_element_by_xpath('/html/body/div/div/form/div[6]/div[3]/ul/li[8]/input').click()
        time.sleep(1.5)
        url=driver.current_url
        print(url)
        i=0
        continue
        #print(2)
        #browser_from=driver.find_elements_by_id('ID_dispDetailBtn')
        #browser_from[i].click()
    time.sleep(0.2)   
    # ウィンドウハンドルを取得する
    #handle_array = driver.window_handles
    # seleniumで操作可能なdriverを切り替える
    #driver.switch_to.window(handle_array[1])
    #try:
     #   WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, 'top_header')))
      #  time.sleep(0.1)
    #except TimeoutException:
     #   pass
        
    i=i+1
    try:
        syamei=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > div:nth-child(4) > table > tbody > tr:nth-child(1) > td:nth-child(1) > table > tbody > tr:nth-child(2) > td').text
        
    except NoSuchElementException:
        syamei='記載なし'
    
    try:
        yubin=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > div:nth-child(4) > table > tbody > tr:nth-child(1) > td:nth-child(1) > table > tbody > tr:nth-child(5) > td:nth-child(2)').text
        
    except NoSuchElementException:
        yubin='記載なし'
    try:    
        jusyo=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > div:nth-child(4) > table > tbody > tr:nth-child(1) > td:nth-child(1) > table > tbody > tr:nth-child(7) > td:nth-child(2)').text
 #   except NoSuchElementException:  
   #     jusyo=driver.find_element_by_id('ID_shgBsJusho').text
    except NoSuchElementException:
        jusyo='記載なし'
    try:
        tel=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > div:nth-child(4) > table > tbody > tr:nth-child(1) > td:nth-child(1) > table > tbody > tr:nth-child(8) > td:nth-child(2)').text
    except NoSuchElementException:
        tel='記載なし'
        
    try:
        fax=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > fieldset > div > table > tbody > tr:nth-child(4) > td:nth-child(2)').text
    except NoSuchElementException:
        fax='記載なし'
        
    try:
        url=driver.find_element_by_css_selector('#page > form > div > div.container-fluid > fieldset > div > table > tbody > tr:nth-child(6) > td:nth-child(2)').text
    except NoSuchElementException:
        url='記載なし'
        
    try:    
        syatyo=driver.find_element_by_id('ID_dhshaMei').text
    except NoSuchElementException:
        syatyo='記載なし'
        
    try:
        tanto=driver.find_element_by_id('ID_ttsTts').text
    except NoSuchElementException:
        tanto='記載なし'

    print(syamei)
    print(yubin)
    print(jusyo)
    print(tel)
    print(fax)
    print(url)
    print(syatyo)
    print(tanto)
    syameilists.append(syamei)
    yubinlists.append(yubin)
    jusyolists.append(jusyo)
    tellists.append(tel)
    faxlists.append(fax)
    urllists.append(url)
    syatyolists.append(syatyo)
    tantolists.append(tanto)
    driver.close()
    # ウィンドウハンドルを取得する
    handle_array = driver.window_handles
    # seleniumで操作可能なdriverを切り替える
    driver.switch_to.window(handle_array[0])


# In[36]:


#エクセルに出力
wb = openpyxl.load_workbook('/Users/naoki/Desktop/Mypandas/案件/ハローワーク_20210220/7製造、修理、保全ー自動車組立、輸送用機械組立/データ入れ.xlsx')
ws = wb["シート1"]
for i in range(0,len(syameilists)):
    ws.cell(row=i+2,column=1,value=syameilists[i])
    # A列
    ws.cell(row=i+2,column=2,value=yubinlists[i])
    # B列
    ws.cell(row=i+2,column=3,value=jusyolists[i])
    ws.cell(row=i+2,column=4,value=tellists[i])
    ws.cell(row=i+2,column=5,value=faxlists[i])
    ws.cell(row=i+2,column=6,value=urllists[i])
    ws.cell(row=i+2,column=7,value=syatyolists[i])
    ws.cell(row=i+2,column=8,value=tantolists[i])
    #ws.cell(row=i+3,column=10,value=delibarylists[i])
wb.save('/Users/naoki/Desktop/Mypandas/案件/ハローワーク_20210220/7製造、修理、保全ー自動車組立、輸送用機械組立/データ入れ.xlsx')


# In[151]:





# In[105]:





# In[ ]:




